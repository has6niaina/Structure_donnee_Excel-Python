import openpyxl

w01 = openpyxl.load_workbook("octobre.xlsx")
w02 = openpyxl.load_workbook("novembre.xlsx")
w03 = openpyxl.load_workbook("decembre.xlsx")
sheet1 = w01.active
donnees= {}
for row in range(2,sheet1.max_row):
	v= sheet1.cell(row, 1).value
	if not v: 
		break
	donnees[v] 
	print(v)
#{"Pommes" : (768,668,980)}